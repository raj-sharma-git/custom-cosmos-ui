import os
import io
import csv
import json
import uuid
import time
import queue
import random
import threading
import traceback
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import (
    Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, Blueprint
)
from flask_session import Session
from azure.identity import ClientSecretCredential
from azure.cosmos import CosmosClient, PartitionKey, exceptions as cosmos_exceptions
from openpyxl import Workbook, load_workbook
from werkzeug.middleware.proxy_fix import ProxyFix

app = Flask(__name__, static_url_path="/cosmos-ui/static")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Server-side Session Configuration
SESSION_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "flask_sessions")
os.makedirs(SESSION_DIR, exist_ok=True)

TEMP_UPLOAD_DIR = os.path.join(SESSION_DIR, "temp_imports")
os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)

TEMP_EXPORT_DIR = os.path.join(SESSION_DIR, "temp_exports")
os.makedirs(TEMP_EXPORT_DIR, exist_ok=True)

app.secret_key = os.environ.get("FLASK_SECRET_KEY", str(uuid.uuid4()))
app.config.update(
    SESSION_TYPE="filesystem",
    SESSION_FILE_DIR=SESSION_DIR,
    SESSION_PERMANENT=True,
    PERMANENT_SESSION_LIFETIME=86400, # 24 hours
)
Session(app)

# In-memory store for active CosmosClient instances
CLIENT_STORE = {}

# Background Ingestion Tasks Store
IMPORT_TASKS = {}
IMPORT_TASKS_LOCK = threading.Lock()

# Background Export Tasks Store
EXPORT_TASKS = {}
EXPORT_TASKS_LOCK = threading.Lock()

# ---------- Blueprint ----------
ui = Blueprint("ui", __name__, url_prefix="/cosmos-ui")

def get_store():
    sid = session.get("sid")
    return CLIENT_STORE.get(sid) if sid else None

def login_required(f):
    from functools import wraps
    @wraps(f)
    def inner(*a, **kw):
        if not get_store():
            flash("Please login first to access this resource.", "warning")
            return redirect(url_for("ui.login"))
        return f(*a, **kw)
    return inner

# ---------- Helper Functions ----------
def execute_with_429_retry(func, *args, max_retries=10, initial_delay=0.1, **kwargs):
    """
    Executes a Cosmos DB SDK operation with automatic HTTP 429 exponential backoff retry.
    Returns: (result, retries_count)
    """
    delay = initial_delay
    retries = 0
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs), retries
        except cosmos_exceptions.CosmosHttpResponseError as e:
            if e.status_code == 429: # RequestRateTooLarge
                retries += 1
                retry_after_ms = e.headers.get("x-ms-retry-after-ms") if hasattr(e, "headers") and e.headers else None
                if retry_after_ms:
                    try:
                        sleep_time = (float(retry_after_ms) / 1000.0) + random.uniform(0.01, 0.05)
                    except Exception:
                        sleep_time = delay + random.uniform(0.01, 0.05)
                else:
                    sleep_time = delay + random.uniform(0.01, 0.05)
                    delay = min(delay * 2, 5.0)
                time.sleep(sleep_time)
            else:
                raise e
    # Final attempt
    return func(*args, **kwargs), retries

def get_partition_key_path(container):
    """Programmatically fetch the partition key path for a container."""
    try:
        properties = container.read()
        paths = properties.get("partitionKey", {}).get("paths", [])
        return paths[0] if paths else None
    except Exception as e:
        print(f"Error fetching partition key path: {e}")
        return "/id" # fallback

def extract_cosmos_error_message(err):
    """Extract a readable, clean message from Azure Cosmos DB SDK exceptions."""
    err_str = str(err)
    try:
        import re
        json_matches = re.findall(r'(\{[^{}]*"errors"[^{}]*\})', err_str, re.DOTALL)
        if not json_matches:
            json_matches = re.findall(r'(\{.*"errors".*\})', err_str, re.DOTALL)
        for jm in json_matches:
            try:
                data = json.loads(jm)
                errors = data.get("errors", [])
                if errors and isinstance(errors, list):
                    first = errors[0]
                    code = first.get("code", "")
                    msg = first.get("message", "")
                    if code and msg:
                        return f"[{code}] {msg}"
                    elif msg:
                        return msg
            except Exception:
                continue

        msg_match = re.search(r'"message":\s*"([^"]+)"', err_str)
        if msg_match:
            code_match = re.search(r'"code":\s*"([^"]+)"', err_str)
            if code_match:
                return f"[{code_match.group(1)}] {msg_match.group(1)}"
            return msg_match.group(1)
    except Exception:
        pass

    if "Content:" in err_str:
        err_str = err_str.split("Content:")[0].strip()
    return err_str.strip()

def extract_partition_key_value(item, pk_path):
    """Dynamically extract the partition key value from an item's attributes based on the path."""
    if not pk_path:
        return None
    parts = pk_path.strip("/").split("/")
    val = item
    for p in parts:
        if isinstance(val, dict) and p in val:
            val = val[p]
        else:
            return None
    return val

def parse_and_build_query(search_mode, search_query, offset=0, limit=10):
    """
    Processes user search input into valid Cosmos DB SQL.
    Supports:
    1. Simple ID search (CONTAINS)
    2. Shorthand WHERE clauses (e.g. c.status = 'active')
    3. Full Cosmos SQL queries (e.g. SELECT c.id, c.name FROM c WHERE ... ORDER BY ...)
    4. Complex queries (GROUP BY, VALUE COUNT(1), TOP ...)
    """
    if not search_query or not search_query.strip():
        return {
            "items_sql": f"SELECT * FROM c ORDER BY c.id OFFSET {offset} LIMIT {limit}",
            "count_sql": "SELECT VALUE COUNT(1) FROM c",
            "params": [],
            "is_custom_projection": False,
            "is_direct_query": False,
            "can_paginate": True
        }

    query = search_query.strip()

    if search_mode == "simple":
        where = "WHERE CONTAINS(LOWER(c.id), @search)"
        params = [{"name": "@search", "value": query.lower()}]
        return {
            "items_sql": f"SELECT * FROM c {where} ORDER BY c.id OFFSET {offset} LIMIT {limit}",
            "count_sql": f"SELECT VALUE COUNT(1) FROM c {where}",
            "params": params,
            "is_custom_projection": False,
            "is_direct_query": False,
            "can_paginate": True
        }

    # Advanced Mode
    upper_q = query.upper()

    # Check if it's a full SELECT query
    if upper_q.startswith("SELECT"):
        has_group_by = "GROUP BY" in upper_q
        has_value = "SELECT VALUE" in upper_q
        has_top = "SELECT TOP" in upper_q
        has_offset_limit = "OFFSET" in upper_q and "LIMIT" in upper_q

        if has_group_by or has_value or has_top or has_offset_limit:
            return {
                "items_sql": query,
                "count_sql": None,
                "params": [],
                "is_custom_projection": not upper_q.startswith("SELECT * FROM C"),
                "is_direct_query": True,
                "can_paginate": False
            }

        # For general SELECT ... FROM c [WHERE ...] [ORDER BY ...]
        count_sql = None
        from_idx = upper_q.find("FROM")
        if from_idx != -1:
            order_idx = upper_q.rfind("ORDER BY")
            if order_idx != -1 and order_idx > from_idx:
                from_where_part = query[from_idx:order_idx].strip()
            else:
                from_where_part = query[from_idx:].strip()
            count_sql = f"SELECT VALUE COUNT(1) {from_where_part}"

        if "ORDER BY" in upper_q:
            items_sql = f"{query} OFFSET {offset} LIMIT {limit}"
        else:
            items_sql = f"{query} ORDER BY c.id OFFSET {offset} LIMIT {limit}"

        return {
            "items_sql": items_sql,
            "count_sql": count_sql,
            "params": [],
            "is_custom_projection": not (upper_q.startswith("SELECT * FROM C") or upper_q.startswith("SELECT * FROM C ")),
            "is_direct_query": False,
            "can_paginate": True
        }
    else:
        # Shorthand WHERE clause
        if upper_q.startswith("WHERE"):
            clean_where = query[5:].strip()
        else:
            clean_where = query.strip()

        return {
            "items_sql": f"SELECT * FROM c WHERE {clean_where} ORDER BY c.id OFFSET {offset} LIMIT {limit}",
            "count_sql": f"SELECT VALUE COUNT(1) FROM c WHERE {clean_where}",
            "params": [],
            "is_custom_projection": False,
            "is_direct_query": False,
            "can_paginate": True
        }

# Legacy helper for export compatibility
def build_search_query(search_mode, search_query):
    meta = parse_and_build_query(search_mode, search_query)
    if search_mode == "advanced":
        if search_query.strip().upper().startswith("SELECT"):
            return "", []
        clean = search_query.strip()
        if clean.upper().startswith("WHERE"):
            clean = clean[5:].strip()
        return f"WHERE {clean}" if clean else "", []
    elif search_mode == "simple" and search_query.strip():
        return "WHERE CONTAINS(LOWER(c.id), @search)", [{"name": "@search", "value": search_query.strip().lower()}]
    return "", []

# ---------- Root Redirect ----------
@app.route("/")
def index_redirect():
    return redirect(url_for("ui.dashboard"))

# ---------- Routes ----------

@ui.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        auth_method = request.form.get("auth_method")
        endpoint = request.form.get("endpoint", "").strip()
        
        # Connection values
        conn_string = request.form.get("conn_string", "").strip()
        account_key = request.form.get("account_key", "").strip()
        
        # Service principal values
        tenant_id = request.form.get("tenant_id", "").strip()
        client_id = request.form.get("client_id", "").strip()
        client_secret = request.form.get("client_secret", "").strip()

        try:
            client = None
            auth_info = {}

            if auth_method == "conn_str":
                if not conn_string:
                    raise ValueError("Connection string is required.")
                client = CosmosClient.from_connection_string(conn_string)
                # Parse endpoint from conn_str for display
                for part in conn_string.split(";"):
                    if part.startswith("AccountEndpoint="):
                        auth_info["endpoint"] = part.replace("AccountEndpoint=", "")
                auth_info["method"] = "Connection String"
            
            elif auth_method == "key":
                if not endpoint or not account_key:
                    raise ValueError("Endpoint URI and Account Key are required.")
                client = CosmosClient(endpoint, credential=account_key, connection_verify=True)
                auth_info["endpoint"] = endpoint
                auth_info["method"] = "Account Key"
                
            elif auth_method == "sp":
                if not endpoint or not tenant_id or not client_id or not client_secret:
                    raise ValueError("All Service Principal fields are required.")
                credential = ClientSecretCredential(tenant_id, client_id, client_secret)
                client = CosmosClient(endpoint, credential=credential, connection_verify=True)
                auth_info["endpoint"] = endpoint
                auth_info["method"] = "Service Principal"
            else:
                raise ValueError("Invalid authentication method selected.")

            # Test connection by listing databases
            databases = list(client.list_databases())
            
            # Setup session
            sid = str(uuid.uuid4())
            session["sid"] = sid
            session["auth_info"] = auth_info
            
            # Store in CLIENT_STORE
            CLIENT_STORE[sid] = {
                "client": client,
                "login_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            flash("Successfully connected to Cosmos DB!", "success")
            return redirect(url_for("ui.dashboard"))

        except Exception as e:
            traceback.print_exc()
            flash(f"Connection failed: {str(e)}", "danger")
            return render_template("login.html")

    return render_template("login.html")

@ui.route("/logout")
def logout():
    sid = session.get("sid")
    if sid in CLIENT_STORE:
        del CLIENT_STORE[sid]
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for("ui.login"))

@ui.route("/")
@ui.route("/dashboard")
@login_required
def dashboard():
    store = get_store()
    client = store["client"]
    try:
        databases = list(client.list_databases())
        db_tree = []
        for db in databases:
            db_client = client.get_database_client(db["id"])
            containers = list(db_client.list_containers())
            db_tree.append({
                "id": db["id"],
                "containers": [c["id"] for c in containers]
            })
        return render_template("dashboard.html", db_tree=db_tree, auth_info=session.get("auth_info"))
    except Exception as e:
        flash(f"Error fetching databases: {str(e)}", "danger")
        return render_template("dashboard.html", db_tree=[], auth_info=session.get("auth_info"))

# ---------- Streaming File Helper for Large Ingestions ----------
def stream_file_records(file_path, filename):
    """
    Generator yielding records from JSON, JSONL, NDJSON, CSV, or XLSX files without high memory usage.
    Yields: (row_index, document_dict)
    """
    fn = filename.lower()
    
    if fn.endswith(".jsonl") or fn.endswith(".ndjson"):
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            for idx, line in enumerate(f):
                line = line.strip()
                if not line:
                    continue
                try:
                    doc = json.loads(line)
                    if isinstance(doc, dict):
                        yield idx + 1, doc
                    else:
                        yield idx + 1, {"_raw_value": doc}
                except Exception as e:
                    yield idx + 1, {"_parse_error": str(e)}

    elif fn.endswith(".json"):
        # Check if JSON array or JSON Lines format
        first_char = ""
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            for char in f.read(1024):
                if not char.isspace():
                    first_char = char
                    break
        
        if first_char == "[":
            # Standard JSON Array
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                data = json.load(f)
                if isinstance(data, list):
                    for idx, doc in enumerate(data):
                        if isinstance(doc, dict):
                            yield idx + 1, doc
                        else:
                            yield idx + 1, {"_raw_value": doc}
                elif isinstance(data, dict):
                    yield 1, data
        else:
            # NDJSON / JSONL
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                for idx, line in enumerate(f):
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        doc = json.loads(line)
                        if isinstance(doc, dict):
                            yield idx + 1, doc
                        else:
                            yield idx + 1, {"_raw_value": doc}
                    except Exception as e:
                        yield idx + 1, {"_parse_error": str(e)}

    elif fn.endswith(".csv"):
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if headers:
                clean_headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers)]
                for idx, row in enumerate(reader):
                    doc = {}
                    for h_idx, h in enumerate(clean_headers):
                        if h_idx < len(row) and row[h_idx] is not None:
                            val = row[h_idx]
                            if isinstance(val, str) and (val.startswith("{") or val.startswith("[")):
                                try:
                                    val = json.loads(val)
                                except Exception:
                                    pass
                            doc[h] = val
                    yield idx + 2, doc

    elif fn.endswith(".xlsx"):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if headers:
            clean_headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(headers)]
            for idx, row in enumerate(rows_iter):
                if not row or all(v is None for v in row):
                    continue
                doc = {}
                for h_idx, h in enumerate(clean_headers):
                    if h_idx < len(row) and row[h_idx] is not None:
                        val = row[h_idx]
                        if isinstance(val, str) and (val.startswith("{") or val.startswith("[")):
                            try:
                                val = json.loads(val)
                            except Exception:
                                pass
                        doc[h] = val
                yield idx + 2, doc
        wb.close()


def run_bulk_import_worker(task_id, file_path, filename, db_id, container_id, pk_path, concurrency, client):
    """
    Ultra-high-throughput continuous streaming worker pool for maximum Cosmos DB ingestion speed.
    Eliminates pipeline stalls by using a continuous thread-safe producer-consumer queue.
    """
    clean_pk = pk_path.strip("/") if pk_path else "id"
    
    with IMPORT_TASKS_LOCK:
        task = IMPORT_TASKS.get(task_id)
        if not task:
            return
        task["status"] = "in_progress"
        task["start_time"] = time.time()

    try:
        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)

        # Scale concurrency up to 200 workers
        num_workers = max(10, min(concurrency, 200))
        doc_queue = queue.Queue(maxsize=num_workers * 25)
        
        stats_lock = threading.Lock()
        total_processed = 0
        total_success = 0
        total_failed = 0
        total_429_retries = 0
        errors = []
        stop_event = threading.Event()

        def worker_loop():
            nonlocal total_processed, total_success, total_failed, total_429_retries
            while not stop_event.is_set():
                try:
                    item = doc_queue.get(timeout=0.15)
                except queue.Empty:
                    continue

                if item is None:
                    doc_queue.task_done()
                    break

                row_idx, doc = item
                if "_parse_error" in doc:
                    with stats_lock:
                        total_processed += 1
                        total_failed += 1
                        if len(errors) < 25:
                            errors.append(f"Row {row_idx}: Parse error: {doc['_parse_error']}")
                    doc_queue.task_done()
                    continue

                if "id" not in doc or not str(doc["id"]).strip():
                    doc["id"] = str(uuid.uuid4())
                else:
                    doc["id"] = str(doc["id"])

                pk_val = extract_partition_key_value(doc, pk_path)
                if pk_val is None:
                    doc[clean_pk] = "imported"

                try:
                    _, retries = execute_with_429_retry(container.upsert_item, body=doc, max_retries=10)
                    with stats_lock:
                        total_processed += 1
                        total_success += 1
                        total_429_retries += retries
                except Exception as err:
                    with stats_lock:
                        total_processed += 1
                        total_failed += 1
                        if len(errors) < 25:
                            errors.append(f"Row {row_idx}: {str(err)}")

                doc_queue.task_done()

        # Start consumer worker threads
        workers = []
        for _ in range(num_workers):
            t = threading.Thread(target=worker_loop, daemon=True)
            t.start()
            workers.append(t)

        # Background metric updater for smooth UI reporting
        def metric_updater():
            while not stop_event.is_set():
                time.sleep(0.2)
                with stats_lock:
                    proc = total_processed
                    succ = total_success
                    fail = total_failed
                    r429 = total_429_retries
                    errs = list(errors)
                elapsed = max(0.1, time.time() - task["start_time"])
                speed = round(proc / elapsed, 1)
                with IMPORT_TASKS_LOCK:
                    task["processed"] = proc
                    task["successful"] = succ
                    task["failed"] = fail
                    task["retries_429"] = r429
                    task["speed_per_sec"] = speed
                    task["errors"] = errs

        metric_thread = threading.Thread(target=metric_updater, daemon=True)
        metric_thread.start()

        # Producer: Stream records from file directly into queue
        records_gen = stream_file_records(file_path, filename)
        for row_info in records_gen:
            with IMPORT_TASKS_LOCK:
                if task.get("cancel_requested"):
                    stop_event.set()
                    break
            while not stop_event.is_set():
                try:
                    doc_queue.put(row_info, timeout=0.1)
                    break
                except queue.Full:
                    with IMPORT_TASKS_LOCK:
                        if task.get("cancel_requested"):
                            stop_event.set()
                            break

        # If cancelled, drain the remaining queue immediately
        if stop_event.is_set():
            while not doc_queue.empty():
                try:
                    doc_queue.get_nowait()
                    doc_queue.task_done()
                except Exception:
                    break

        # Send termination sentinels
        for _ in range(num_workers):
            try:
                doc_queue.put_nowait(None)
            except Exception:
                pass

        # Await workers with small timeout
        for t in workers:
            t.join(timeout=0.2)

        stop_event.set()
        metric_thread.join(timeout=0.3)

        elapsed = max(0.1, time.time() - task["start_time"])
        speed = round(total_processed / elapsed, 1)

        with IMPORT_TASKS_LOCK:
            if task.get("cancel_requested"):
                task["status"] = "cancelled"
            elif task.get("status") != "cancelled":
                task["status"] = "completed"
            task["end_time"] = time.time()
            task["processed"] = total_processed
            task["successful"] = total_success
            task["failed"] = total_failed
            task["retries_429"] = total_429_retries
            task["speed_per_sec"] = speed
            task["errors"] = errors

    except Exception as e:
        traceback.print_exc()
        with IMPORT_TASKS_LOCK:
            task["status"] = "failed"
            task["error_message"] = str(e)
            task["end_time"] = time.time()
            
    finally:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass


def run_export_worker(task_id, file_path, filename, format_type, db_id, container_id, query_sql, query_params, client):
    """
    Background worker that streams documents from Cosmos DB directly into disk
    supporting JSON, JSONL, CSV, and XLSX with 429 rate limit backoff and live progress.
    """
    with EXPORT_TASKS_LOCK:
        task = EXPORT_TASKS.get(task_id)
        if not task:
            return
        task["status"] = "in_progress"
        task["start_time"] = time.time()

    try:
        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)

        processed = 0
        total_429_retries = 0

        # Query Cosmos DB with cross-partition support and 1,000-item page streaming
        query_iterable = container.query_items(
            query=query_sql,
            parameters=query_params,
            enable_cross_partition_query=True,
            max_item_count=1000
        )
        pager = query_iterable.by_page()

        def stream_items_with_retry():
            nonlocal total_429_retries
            while True:
                with EXPORT_TASKS_LOCK:
                    if task.get("cancel_requested"):
                        return

                def fetch_page():
                    try:
                        return next(pager, None)
                    except StopIteration:
                        return None

                page, retries = execute_with_429_retry(fetch_page, max_retries=10)
                total_429_retries += retries
                if page is None:
                    break
                for item in page:
                    yield item

        # Stream directly into target file format without high RAM consumption
        if format_type in ["jsonl", "ndjson"]:
            with open(file_path, "w", encoding="utf-8") as f:
                for doc in stream_items_with_retry():
                    with EXPORT_TASKS_LOCK:
                        if task.get("cancel_requested"):
                            break
                    f.write(json.dumps(doc) + "\n")
                    processed += 1
                    if processed % 200 == 0:
                        elapsed = max(0.1, time.time() - task["start_time"])
                        with EXPORT_TASKS_LOCK:
                            task["processed"] = processed
                            task["speed_per_sec"] = round(processed / elapsed, 1)
                            task["retries_429"] = total_429_retries

        elif format_type == "json":
            with open(file_path, "w", encoding="utf-8") as f:
                f.write("[\n")
                first = True
                for doc in stream_items_with_retry():
                    with EXPORT_TASKS_LOCK:
                        if task.get("cancel_requested"):
                            break
                    if not first:
                        f.write(",\n")
                    else:
                        first = False
                    f.write(json.dumps(doc, indent=2))
                    processed += 1
                    if processed % 200 == 0:
                        elapsed = max(0.1, time.time() - task["start_time"])
                        with EXPORT_TASKS_LOCK:
                            task["processed"] = processed
                            task["speed_per_sec"] = round(processed / elapsed, 1)
                            task["retries_429"] = total_429_retries
                f.write("\n]\n")

        elif format_type == "csv":
            with open(file_path, "w", encoding="utf-8", newline="") as f:
                writer = None
                headers = []
                for doc in stream_items_with_retry():
                    with EXPORT_TASKS_LOCK:
                        if task.get("cancel_requested"):
                            break
                    if writer is None:
                        if isinstance(doc, dict):
                            headers = sorted(list(doc.keys()))
                            if "id" in headers:
                                headers.remove("id")
                                headers.insert(0, "id")
                        else:
                            headers = ["value"]
                        writer = csv.writer(f)
                        writer.writerow(headers)

                    if isinstance(doc, dict):
                        row = []
                        for h in headers:
                            val = doc.get(h, "")
                            if isinstance(val, (dict, list)):
                                val = json.dumps(val)
                            row.append(val)
                        writer.writerow(row)
                    else:
                        writer.writerow([doc])

                    processed += 1
                    if processed % 200 == 0:
                        elapsed = max(0.1, time.time() - task["start_time"])
                        with EXPORT_TASKS_LOCK:
                            task["processed"] = processed
                            task["speed_per_sec"] = round(processed / elapsed, 1)
                            task["retries_429"] = total_429_retries

        elif format_type == "xlsx":
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="CosmosExport")
            headers_written = False
            headers = []
            for doc in stream_items_with_retry():
                with EXPORT_TASKS_LOCK:
                    if task.get("cancel_requested"):
                        break
                if not headers_written:
                    if isinstance(doc, dict):
                        headers = sorted(list(doc.keys()))
                        if "id" in headers:
                            headers.remove("id")
                            headers.insert(0, "id")
                    else:
                        headers = ["value"]
                    ws.append(headers)
                    headers_written = True

                if isinstance(doc, dict):
                    row = []
                    for h in headers:
                        val = doc.get(h, "")
                        if isinstance(val, (dict, list)):
                            val = json.dumps(val)
                        row.append(val)
                    ws.append(row)
                else:
                    ws.append([doc])

                processed += 1
                if processed % 200 == 0:
                    elapsed = max(0.1, time.time() - task["start_time"])
                    with EXPORT_TASKS_LOCK:
                        task["processed"] = processed
                        task["speed_per_sec"] = round(processed / elapsed, 1)
                        task["retries_429"] = total_429_retries
            wb.save(file_path)

        elapsed = max(0.1, time.time() - task["start_time"])
        with EXPORT_TASKS_LOCK:
            if task.get("cancel_requested"):
                task["status"] = "cancelled"
            else:
                task["status"] = "completed"
            task["end_time"] = time.time()
            task["processed"] = processed
            task["speed_per_sec"] = round(processed / elapsed, 1)
            task["retries_429"] = total_429_retries

    except Exception as e:
        traceback.print_exc()
        with EXPORT_TASKS_LOCK:
            task["status"] = "failed"
            task["error_message"] = str(e)
            task["end_time"] = time.time()


@ui.route("/db/<db_id>/container/<container_id>")
@login_required
def container_view(db_id, container_id):
    store = get_store()
    client = store["client"]
    
    # Query parameters
    page = request.args.get("page", 1, type=int)
    limit = request.args.get("limit", 10, type=int)
    search_mode = request.args.get("search_mode", "simple")
    search_query = request.args.get("search_query", "")
    cached_total_str = request.args.get("total_items", None)

    offset = (page - 1) * limit

    try:
        # Get DB and Container Clients
        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)
        
        # Programmatically detect Partition Key
        pk_path = get_partition_key_path(container)
        
        # Build query parts using upgraded parser
        q_meta = parse_and_build_query(search_mode, search_query, offset=offset, limit=limit)
        
        # Determine total_items (Smart Count for 20L+ records)
        total_items = None
        if q_meta["is_direct_query"]:
            total_items = None # Direct custom queries might return arbitrary rows
        elif cached_total_str is not None and str(cached_total_str).strip() != "":
            try:
                total_items = int(cached_total_str)
            except ValueError:
                total_items = None
                
        if total_items is None and q_meta["count_sql"]:
            try:
                count_iter = container.query_items(
                    query=q_meta["count_sql"],
                    parameters=q_meta["params"],
                    enable_cross_partition_query=True
                )
                count_res = list(count_iter)
                total_items = int(count_res[0]) if count_res and count_res[0] is not None else 0
            except Exception as count_err:
                print(f"Count query notice (e.g. large dataset scan): {count_err}")
                total_items = 0

        query_error = None
        raw_items = []

        # Execute items query with graceful error handling
        try:
            items_iter = container.query_items(
                query=q_meta["items_sql"],
                parameters=q_meta["params"],
                enable_cross_partition_query=True,
                max_item_count=limit
            )
            raw_items = list(items_iter)
        except Exception as q_err:
            traceback.print_exc()
            query_error = extract_cosmos_error_message(q_err)
            raw_items = []
        
        if query_error:
            total_items = 0
        elif total_items is None or total_items == 0:
            if q_meta["is_direct_query"]:
                total_items = len(raw_items)
            elif not q_meta["count_sql"]:
                total_items = len(raw_items)

        # Process items to extract partition key value, summary, or custom projections
        processed_items = []
        custom_columns = []
        all_keys = set()

        for it in raw_items:
            if isinstance(it, dict):
                has_id = "id" in it
                pk_val = extract_partition_key_value(it, pk_path) if has_id else None
                all_keys.update(it.keys())
                processed_items.append({
                    "id": it.get("id", None),
                    "pk_val": pk_val,
                    "has_id": has_id,
                    "raw": it
                })
            else:
                processed_items.append({
                    "id": None,
                    "pk_val": None,
                    "has_id": False,
                    "raw": {"_result": it}
                })

        # If custom projection without standard id, determine custom columns to display
        is_custom_projection = q_meta["is_custom_projection"] or (len(processed_items) > 0 and not any(p["has_id"] for p in processed_items))
        if is_custom_projection and all_keys:
            custom_columns = sorted([k for k in all_keys if not k.startswith("_")])[:8]
            if not custom_columns:
                custom_columns = sorted(list(all_keys))[:8]

        # Calculate pages
        effective_total = total_items if total_items is not None else len(processed_items)
        total_pages = max(1, (effective_total + limit - 1) // limit)

        # Database tree for sidebar quick-nav
        databases = list(client.list_databases())
        db_tree = []
        for db in databases:
            dbc = client.get_database_client(db["id"])
            db_tree.append({
                "id": db["id"],
                "containers": [c["id"] for c in dbc.list_containers()]
            })

        return render_template(
            "container.html",
            db_id=db_id,
            container_id=container_id,
            pk_path=pk_path,
            items=processed_items,
            page=page,
            limit=limit,
            search_mode=search_mode,
            search_query=search_query,
            total_items=effective_total,
            total_pages=total_pages,
            is_custom_projection=is_custom_projection,
            custom_columns=custom_columns,
            can_paginate=q_meta["can_paginate"],
            query_error=query_error,
            db_tree=db_tree,
            auth_info=session.get("auth_info")
        )
    except Exception as e:
        traceback.print_exc()
        flash(f"Error accessing container: {extract_cosmos_error_message(e)}", "danger")
        return redirect(url_for("ui.dashboard"))

# ---------- API Endpoints ----------

@ui.route("/api/db/<db_id>/container/<container_id>/item", methods=["POST"])
@login_required
def api_upsert_item(db_id, container_id):
    """Create or Update an item in Cosmos DB (Upsert)"""
    store = get_store()
    client = store["client"]
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "No data provided"}), 400
        
        if "id" not in data or not str(data["id"]).strip():
            return jsonify({"status": "error", "message": "Document must contain an 'id' attribute."}), 400
            
        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)
        
        # Verify partition key exists in document
        pk_path = get_partition_key_path(container)
        pk_val = extract_partition_key_value(data, pk_path)
        if pk_val is None:
            clean_pk = pk_path.strip("/")
            return jsonify({
                "status": "error", 
                "message": f"Document must contain the partition key path attribute: '{clean_pk}'"
            }), 400
            
        # Execute Upsert with 429 retry
        res, _ = execute_with_429_retry(container.upsert_item, body=data)
        return jsonify({"status": "success", "message": "Document saved successfully", "item": res})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@ui.route("/api/db/<db_id>/container/<container_id>/item/delete", methods=["POST"])
@login_required
def api_delete_item(db_id, container_id):
    """Delete an item from Cosmos DB"""
    store = get_store()
    client = store["client"]
    
    try:
        data = request.get_json()
        item_id = data.get("id")
        partition_key = data.get("partition_key")
        
        if not item_id:
            return jsonify({"status": "error", "message": "Item ID is required."}), 400

        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)
        
        # Delete item with 429 retry
        execute_with_429_retry(container.delete_item, item=item_id, partition_key=partition_key)
        return jsonify({"status": "success", "message": "Document deleted successfully."})

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@ui.route("/api/db/<db_id>/container/<container_id>/items/bulk-delete", methods=["POST"])
@login_required
def api_bulk_delete_items(db_id, container_id):
    """Bulk delete a list of items from Cosmos DB with 429 rate limit backoff"""
    store = get_store()
    client = store["client"]
    try:
        data = request.get_json()
        if not data or "items" not in data:
            return jsonify({"status": "error", "message": "No items provided for deletion."}), 400
            
        items_to_delete = data.get("items", [])
        if not items_to_delete:
            return jsonify({"status": "error", "message": "Item list is empty."}), 400

        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)
        
        deleted_count = 0
        failed_count = 0
        total_retries = 0
        errors = []
        
        def delete_single_item(item_info):
            item_id = item_info.get("id")
            pk_val = item_info.get("partition_key")
            if not item_id:
                return False, 0, "Missing item ID"
            try:
                _, retries = execute_with_429_retry(container.delete_item, item=item_id, partition_key=pk_val)
                return True, retries, None
            except Exception as e:
                return False, 0, f"ID {item_id}: {str(e)}"
                
        max_workers = min(25, max(1, len(items_to_delete)))
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_item = {executor.submit(delete_single_item, item): item for item in items_to_delete}
            for future in as_completed(future_to_item):
                success, retries, err_msg = future.result()
                total_retries += retries
                if success:
                    deleted_count += 1
                else:
                    failed_count += 1
                    if len(errors) < 5 and err_msg:
                        errors.append(err_msg)
                        
        msg = f"Successfully deleted {deleted_count} document(s)."
        if total_retries > 0:
            msg += f" (Handled {total_retries} rate limit retries)"
        if failed_count > 0:
            msg += f" {failed_count} item(s) failed."
            
        return jsonify({
            "status": "success" if deleted_count > 0 else "error",
            "message": msg,
            "deleted_count": deleted_count,
            "failed_count": failed_count,
            "retries_429": total_retries,
            "errors": errors
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@ui.route("/api/db/<db_id>/container/<container_id>/import-async", methods=["POST"])
@login_required
def api_async_import_items(db_id, container_id):
    """Initiates an asynchronous background bulk ingestion task with live progress tracking"""
    store = get_store()
    client = store["client"]
    
    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"status": "error", "message": "No file selected for import."}), 400
        
    filename = file.filename
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ["json", "jsonl", "ndjson", "csv", "xlsx"]:
        return jsonify({"status": "error", "message": "Unsupported format. Please upload JSON, JSONL, NDJSON, CSV, or XLSX."}), 400

    try:
        concurrency = request.form.get("concurrency", 100, type=int)
        concurrency = max(10, min(concurrency, 200))
        
        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)
        pk_path = get_partition_key_path(container)
        
        task_id = str(uuid.uuid4())
        temp_file_path = os.path.join(TEMP_UPLOAD_DIR, f"{task_id}_{filename}")
        file.save(temp_file_path)
        
        # Estimate total records for quick progress estimation
        total_estimate = 0
        try:
            if ext in ["jsonl", "ndjson", "csv"]:
                with open(temp_file_path, "r", encoding="utf-8", errors="replace") as f:
                    for _ in f:
                        total_estimate += 1
                if ext == "csv" and total_estimate > 0:
                    total_estimate -= 1
        except Exception:
            total_estimate = 0

        with IMPORT_TASKS_LOCK:
            IMPORT_TASKS[task_id] = {
                "task_id": task_id,
                "db_id": db_id,
                "container_id": container_id,
                "filename": filename,
                "status": "starting",
                "total_estimate": total_estimate,
                "processed": 0,
                "successful": 0,
                "failed": 0,
                "retries_429": 0,
                "speed_per_sec": 0,
                "start_time": time.time(),
                "end_time": None,
                "errors": [],
                "cancel_requested": False
            }
            
        # Spawn daemon worker thread
        worker_thread = threading.Thread(
            target=run_bulk_import_worker,
            args=(task_id, temp_file_path, filename, db_id, container_id, pk_path, concurrency, client),
            daemon=True
        )
        worker_thread.start()
        
        return jsonify({
            "status": "success",
            "task_id": task_id,
            "message": "Bulk import job initiated successfully.",
            "total_estimate": total_estimate
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Failed to start import task: {str(e)}"}), 500

@ui.route("/api/import-task/<task_id>", methods=["GET"])
@login_required
def api_get_import_task(task_id):
    """Fetch live progress metrics of a background bulk import task"""
    with IMPORT_TASKS_LOCK:
        task = IMPORT_TASKS.get(task_id)
        if not task:
            return jsonify({"status": "error", "message": "Task not found"}), 404
            
        elapsed = (task["end_time"] if task.get("end_time") else time.time()) - task["start_time"]
        
        return jsonify({
            "status": "success",
            "task": {
                "task_id": task["task_id"],
                "status": task["status"],
                "filename": task["filename"],
                "total_estimate": task["total_estimate"],
                "processed": task["processed"],
                "successful": task["successful"],
                "failed": task["failed"],
                "retries_429": task["retries_429"],
                "speed_per_sec": task["speed_per_sec"],
                "elapsed_seconds": round(elapsed, 1),
                "errors": task["errors"][:15],
                "error_message": task.get("error_message", None)
            }
        })

@ui.route("/api/import-task/<task_id>/cancel", methods=["POST"])
@login_required
def api_cancel_import_task(task_id):
    """Requests graceful cancellation of a running bulk import task"""
    with IMPORT_TASKS_LOCK:
        task = IMPORT_TASKS.get(task_id)
        if not task:
            return jsonify({"status": "error", "message": "Task not found"}), 404
        task["cancel_requested"] = True
        task["status"] = "cancelled"
        task["end_time"] = time.time()
        return jsonify({"status": "success", "message": "Cancellation requested."})

@ui.route("/api/db/<db_id>/container/<container_id>/empty", methods=["POST"])
@login_required
def api_empty_container(db_id, container_id):
    """
    Empties all documents from a container by recreating it with identical schema/configuration
    or by bulk-deleting all documents.
    """
    store = get_store()
    client = store["client"]
    try:
        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)
        
        # Read current container properties (partition key, indexing policy, default_ttl, etc.)
        properties = container.read()
        pk_info = properties.get("partitionKey", {})
        pk_paths = pk_info.get("paths", ["/id"])
        pk_path = pk_paths[0] if pk_paths else "/id"
        indexing_policy = properties.get("indexingPolicy")
        default_ttl = properties.get("defaultTtl")
        unique_key_policy = properties.get("uniqueKeyPolicy")
        
        # Method 1: Drop & Recreate container (Instant deletion of millions of documents with 0 RU cost per item)
        try:
            db_client.delete_container(container_id)
            
            create_kwargs = {
                "id": container_id,
                "partition_key": PartitionKey(path=pk_path)
            }
            if indexing_policy:
                create_kwargs["indexing_policy"] = indexing_policy
            if default_ttl is not None:
                create_kwargs["default_ttl"] = default_ttl
            if unique_key_policy:
                create_kwargs["unique_key_policy"] = unique_key_policy
                
            db_client.create_container(**create_kwargs)
            return jsonify({
                "status": "success",
                "message": f"Container '{container_id}' was emptied successfully (recreated with original schema)."
            })
        except Exception as recreate_err:
            print(f"Container recreate notice, falling back to document truncate: {recreate_err}")
            # Method 2: Fallback query and bulk delete
            items = list(container.query_items("SELECT c.id FROM c", enable_cross_partition_query=True))
            deleted_count = 0
            for it in items:
                pk_val = extract_partition_key_value(it, pk_path)
                try:
                    execute_with_429_retry(container.delete_item, item=it["id"], partition_key=pk_val)
                    deleted_count += 1
                except Exception:
                    pass
            return jsonify({
                "status": "success",
                "message": f"Emptied {deleted_count} document(s) from container '{container_id}'."
            })
            
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Failed to empty container: {str(e)}"}), 500

@ui.route("/api/db/<db_id>/container/<container_id>/export-async", methods=["POST"])
@login_required
def api_async_export_items(db_id, container_id):
    """Initiates an asynchronous background export task with live progress tracking"""
    store = get_store()
    client = store["client"]

    try:
        data = request.get_json() or {}
        format_type = data.get("format", "jsonl").lower()
        if format_type not in ["json", "jsonl", "ndjson", "csv", "xlsx"]:
            format_type = "jsonl"

        search_mode = data.get("search_mode", "simple")
        search_query = data.get("search_query", "")

        q_meta = parse_and_build_query(search_mode, search_query, offset=0, limit=10000000)
        export_sql = q_meta["items_sql"]
        # Strip pagination offset/limit to export full dataset
        if "OFFSET" in export_sql.upper():
            offset_pos = export_sql.upper().rfind("OFFSET")
            export_sql = export_sql[:offset_pos].strip()

        task_id = str(uuid.uuid4())
        ext = "xlsx" if format_type == "xlsx" else ("csv" if format_type == "csv" else ("json" if format_type == "json" else "jsonl"))
        filename = f"cosmos_{container_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
        temp_file_path = os.path.join(TEMP_EXPORT_DIR, f"{task_id}_{filename}")

        # Total count estimate
        total_estimate = data.get("total_estimate", 0)

        with EXPORT_TASKS_LOCK:
            EXPORT_TASKS[task_id] = {
                "task_id": task_id,
                "db_id": db_id,
                "container_id": container_id,
                "filename": filename,
                "file_path": temp_file_path,
                "format": format_type,
                "status": "starting",
                "total_estimate": total_estimate,
                "processed": 0,
                "retries_429": 0,
                "speed_per_sec": 0,
                "start_time": time.time(),
                "end_time": None,
                "cancel_requested": False
            }

        worker_thread = threading.Thread(
            target=run_export_worker,
            args=(task_id, temp_file_path, filename, format_type, db_id, container_id, export_sql, q_meta["params"], client),
            daemon=True
        )
        worker_thread.start()

        return jsonify({
            "status": "success",
            "task_id": task_id,
            "message": "Export job initiated successfully.",
            "filename": filename
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": f"Failed to start export: {str(e)}"}), 500


@ui.route("/api/export-task/<task_id>", methods=["GET"])
@login_required
def api_get_export_task(task_id):
    """Fetch live progress metrics of a background export task"""
    with EXPORT_TASKS_LOCK:
        task = EXPORT_TASKS.get(task_id)
        if not task:
            return jsonify({"status": "error", "message": "Task not found"}), 404

        elapsed = (task["end_time"] if task.get("end_time") else time.time()) - task["start_time"]

        return jsonify({
            "status": "success",
            "task": {
                "task_id": task["task_id"],
                "status": task["status"],
                "filename": task["filename"],
                "total_estimate": task["total_estimate"],
                "processed": task["processed"],
                "retries_429": task["retries_429"],
                "speed_per_sec": task["speed_per_sec"],
                "elapsed_seconds": round(elapsed, 1),
                "error_message": task.get("error_message", None)
            }
        })


@ui.route("/api/export-task/<task_id>/cancel", methods=["POST"])
@login_required
def api_cancel_export_task(task_id):
    """Requests cancellation of a running export task"""
    with EXPORT_TASKS_LOCK:
        task = EXPORT_TASKS.get(task_id)
        if not task:
            return jsonify({"status": "error", "message": "Task not found"}), 404
        task["cancel_requested"] = True
        task["status"] = "cancelled"
        task["end_time"] = time.time()
        return jsonify({"status": "success", "message": "Export cancellation requested."})


@ui.route("/api/export-task/<task_id>/download", methods=["GET"])
@login_required
def api_download_export_task(task_id):
    """Download the completed exported file"""
    with EXPORT_TASKS_LOCK:
        task = EXPORT_TASKS.get(task_id)
        if not task or task.get("status") != "completed":
            flash("Export file not ready or task not found.", "warning")
            return redirect(url_for("ui.dashboard"))

        file_path = task["file_path"]
        filename = task["filename"]

    if not os.path.exists(file_path):
        flash("Export file does not exist on server.", "danger")
        return redirect(url_for("ui.dashboard"))

    mimetype = "application/octet-stream"
    if filename.endswith(".json"):
        mimetype = "application/json"
    elif filename.endswith(".jsonl") or filename.endswith(".ndjson"):
        mimetype = "application/x-ndjson"
    elif filename.endswith(".csv"):
        mimetype = "text/csv"
    elif filename.endswith(".xlsx"):
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return send_file(
        file_path,
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename
    )

@ui.route("/db/<db_id>/container/<container_id>/import", methods=["POST"])
@login_required
def import_items(db_id, container_id):
    """Synchronous import fallback handling JSON, JSONL, CSV, and XLSX with 429 retries"""
    store = get_store()
    client = store["client"]
    
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("No file selected for import.", "warning")
        return redirect(url_for("ui.container_view", db_id=db_id, container_id=container_id))

    try:
        filename = file.filename
        temp_path = os.path.join(TEMP_UPLOAD_DIR, f"sync_{uuid.uuid4()}_{filename}")
        file.save(temp_path)
        
        db_client = client.get_database_client(db_id)
        container = db_client.get_container_client(container_id)
        pk_path = get_partition_key_path(container)
        clean_pk = pk_path.strip("/") if pk_path else "id"

        imported_count = 0
        skipped_count = 0
        total_429_retries = 0
        errors = []

        try:
            for idx, doc in stream_file_records(temp_path, filename):
                if "_parse_error" in doc:
                    skipped_count += 1
                    errors.append(f"Row {idx}: {doc['_parse_error']}")
                    continue
                try:
                    if "id" not in doc or not str(doc["id"]).strip():
                        doc["id"] = str(uuid.uuid4())
                    pk_val = extract_partition_key_value(doc, pk_path)
                    if pk_val is None:
                        doc[clean_pk] = "imported"
                    _, retries = execute_with_429_retry(container.upsert_item, body=doc, max_retries=10)
                    total_429_retries += retries
                    imported_count += 1
                except Exception as item_err:
                    skipped_count += 1
                    if len(errors) < 5:
                        errors.append(f"Row {idx}: {str(item_err)}")
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

        msg = f"Import Summary: {imported_count} documents imported successfully."
        if total_429_retries > 0:
            msg += f" (Handled {total_429_retries} rate limit 429 retries)"
        if skipped_count > 0:
            msg += f" {skipped_count} items failed. Sample errors: {', '.join(errors[:3])}"
            flash(msg, "warning")
        else:
            flash(msg, "success")

    except Exception as e:
        traceback.print_exc()
        flash(f"Import process failed: {str(e)}", "danger")

    return redirect(url_for("ui.container_view", db_id=db_id, container_id=container_id))

@ui.route("/api/provision", methods=["POST"])
@login_required
def api_provision():
    """Provision a new Database and optionally a Container inside it"""
    store = get_store()
    client = store["client"]
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "No data provided"}), 400
            
        db_id = data.get("db_id", "").strip()
        container_id = data.get("container_id", "").strip()
        partition_key = data.get("partition_key", "").strip()
        
        if not db_id:
            return jsonify({"status": "error", "message": "Database ID is required."}), 400
            
        existing_dbs = {db["id"] for db in client.list_databases()}
        
        if container_id:
            if db_id in existing_dbs:
                db_client = client.get_database_client(db_id)
                existing_containers = {c["id"] for c in db_client.list_containers()}
                if container_id in existing_containers:
                    return jsonify({"status": "error", "message": f"Container '{container_id}' already exists in database '{db_id}'."}), 400
            
            client.create_database_if_not_exists(id=db_id)
            
            if not partition_key:
                partition_key = "/id"
            if not partition_key.startswith("/"):
                partition_key = "/" + partition_key
                
            db_client = client.get_database_client(db_id)
            db_client.create_container_if_not_exists(
                id=container_id,
                partition_key=PartitionKey(path=partition_key)
            )
            msg = f"Successfully provisioned database '{db_id}' and container '{container_id}'."
        else:
            if db_id in existing_dbs:
                return jsonify({"status": "error", "message": f"Database '{db_id}' already exists."}), 400
                
            client.create_database_if_not_exists(id=db_id)
            msg = f"Successfully created database '{db_id}'."
            
        return jsonify({
            "status": "success", 
            "message": msg
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@ui.route("/api/db/<db_id>/delete", methods=["POST"])
@login_required
def api_delete_database(db_id):
    """Delete a Database from Cosmos DB"""
    store = get_store()
    client = store["client"]
    try:
        client.delete_database(db_id)
        return jsonify({"status": "success", "message": f"Database '{db_id}' deleted successfully."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@ui.route("/api/db/<db_id>/container/<container_id>/delete", methods=["POST"])
@login_required
def api_delete_container(db_id, container_id):
    """Delete a Container from a Database"""
    store = get_store()
    client = store["client"]
    try:
        db_client = client.get_database_client(db_id)
        db_client.delete_container(container_id)
        return jsonify({"status": "success", "message": f"Container '{container_id}' deleted successfully from database '{db_id}'."})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@ui.route("/api/bulk-check", methods=["POST"])
@login_required
def api_bulk_check():
    """Parse bulk upload file and check for existing databases"""
    store = get_store()
    client = store["client"]
    
    file = request.files.get("file")
    if not file or file.filename == "":
        return jsonify({"status": "error", "message": "No file provided"}), 400
        
    try:
        filename = file.filename.lower()
        rows = []
        headers = []
        
        file_bytes = file.stream.read()
        file_io = io.BytesIO(file_bytes)

        if filename.endswith(".csv"):
            stream = io.StringIO(file_bytes.decode("utf-8"), newline=None)
            reader = csv.reader(stream)
            headers = next(reader, None)
            if headers:
                for r in reader:
                    rows.append(r)
        elif filename.endswith(".xlsx"):
            wb = load_workbook(file_io, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if headers:
                for r in rows_iter:
                    rows.append(r)
        else:
            return jsonify({"status": "error", "message": "Unsupported file format"}), 400
            
        if not headers:
            return jsonify({"status": "error", "message": "File is empty or missing headers"}), 400
            
        db_idx = -1
        for h_idx, h in enumerate(headers):
            if h is None:
                continue
            h_str = str(h).strip().lower()
            if h_str in ["db_name", "database_name", "database", "db", "dbname"]:
                db_idx = h_idx
        
        if db_idx == -1 and len(headers) > 0:
            db_idx = 0
            
        uploaded_dbs = set()
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            db_name = str(row[db_idx]).strip() if db_idx < len(row) and row[db_idx] is not None else ""
            if db_name:
                uploaded_dbs.add(db_name)
                
        existing_cosmos_dbs = {db["id"] for db in client.list_databases()}
        overlap = list(uploaded_dbs.intersection(existing_cosmos_dbs))
        
        return jsonify({"status": "success", "existing_dbs": overlap})
        
    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500

@ui.route("/bulk-create", methods=["POST"])
@login_required
def bulk_create_dbs_containers():
    """Bulk provision databases and containers from CSV or Excel (.xlsx) file"""
    store = get_store()
    client = store["client"]
    
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("No file selected for bulk creation.", "warning")
        return redirect(url_for("ui.dashboard"))
        
    try:
        filename = file.filename.lower()
        rows = []
        headers = []
        
        file_bytes = file.stream.read()
        file_io = io.BytesIO(file_bytes)

        if filename.endswith(".csv"):
            stream = io.StringIO(file_bytes.decode("utf-8"), newline=None)
            reader = csv.reader(stream)
            headers = next(reader, None)
            if headers:
                for r in reader:
                    rows.append(r)
        elif filename.endswith(".xlsx"):
            wb = load_workbook(file_io, read_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if headers:
                for r in rows_iter:
                    rows.append(r)
        else:
            flash("Unsupported file format. Please upload a CSV or XLSX file.", "danger")
            return redirect(url_for("ui.dashboard"))
            
        if not headers:
            flash("Uploaded file is empty or missing headers.", "danger")
            return redirect(url_for("ui.dashboard"))
            
        # Standardize headers to find indices
        db_idx, container_idx, pk_idx = -1, -1, -1
        for h_idx, h in enumerate(headers):
            if h is None:
                continue
            h_str = str(h).strip().lower()
            if h_str in ["db_name", "database_name", "database", "db", "dbname"]:
                db_idx = h_idx
            elif h_str in ["container_name", "container", "collection", "containername"]:
                container_idx = h_idx
            elif h_str in ["partition_key", "partition_path", "partitionkey", "pk", "partitionkeypath"]:
                pk_idx = h_idx
                
        # Robust positions fallback if header names aren't matched
        if db_idx == -1 and len(headers) > 0:
            db_idx = 0
        if container_idx == -1 and len(headers) > 1:
            container_idx = 1
        if pk_idx == -1 and len(headers) > 2:
            pk_idx = 2
            
        created_dbs = set()
        created_containers = []
        errors = []
        
        for idx, row in enumerate(rows):
            if not row or all(v is None for v in row):
                continue
                
            try:
                db_name = str(row[db_idx]).strip() if db_idx < len(row) and row[db_idx] is not None else ""
                container_name = str(row[container_idx]).strip() if container_idx < len(row) and row[container_idx] is not None else ""
                partition_key = str(row[pk_idx]).strip() if (pk_idx != -1 and pk_idx < len(row) and row[pk_idx] is not None) else "/id"
                
                if not db_name or not container_name:
                    continue
                    
                mode = request.form.get("mode", "merge")
                
                # Create Database if not exists
                if mode == "overwrite" and db_name not in created_dbs:
                    try:
                        client.get_database_client(db_name).read()
                        client.delete_database(db_name)
                    except exceptions.CosmosResourceNotFoundError:
                        pass

                client.create_database_if_not_exists(id=db_name)
                created_dbs.add(db_name)
                
                # Clean Partition Key
                pk_path = partition_key if partition_key else "/id"
                if not pk_path.startswith("/"):
                    pk_path = "/" + pk_path
                    
                # Create Container if not exists
                db_client = client.get_database_client(db_name)
                db_client.create_container_if_not_exists(
                    id=container_name,
                    partition_key=PartitionKey(path=pk_path)
                )
                created_containers.append(f"{db_name}/{container_name}")
                
            except Exception as row_err:
                errors.append(f"Row {idx+2}: {str(row_err)}")
                
        msg = f"Bulk Process Complete! Verified/Created {len(created_dbs)} databases and {len(created_containers)} containers."
        if errors:
            msg += f" Encountered {len(errors)} errors. Sample errors: {', '.join(errors[:3])}"
            flash(msg, "warning")
        else:
            flash(msg, "success")
            
    except Exception as e:
        traceback.print_exc()
        flash(f"Bulk creation failed: {str(e)}", "danger")
        
    return redirect(url_for("ui.dashboard"))

# Register blueprint
app.register_blueprint(ui)

# ---------- Run App ----------
if __name__ == "__main__":
    # Local dev server runs on 5001 to prevent conflicts
    app.run(host="0.0.0.0", port=5001, debug=True)